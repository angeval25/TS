"""
Script para obtener issues desde Jira usando JQL y actualizar Libro1.xlsx
Ejecuta una consulta JQL y llena la columna Clave en Libro1.xlsx
"""
from jira_integration import JiraIntegration
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime, timezone

def obtener_issues_y_actualizar_xlsx(archivo_xlsx='Libro1.xlsx', max_results=None):
    """
    Obtiene issues desde Jira usando JQL y actualiza Libro1.xlsx con las claves
    
    Args:
        archivo_xlsx: Nombre del archivo XLSX a actualizar
        max_results: Número máximo de resultados a obtener (None para todos)
    """
    
    # Consulta JQL usando horas (-720h = 30 días)
    # Usamos horas en lugar de días para mayor precisión y consistencia
    jql_query = 'created >= -720h AND project = TPGSOC AND assignee IN membersOf("RSOC ILATAM L1") ORDER BY created DESC'
    
    print("=" * 80)
    print("Obteniendo issues desde Jira")
    print("=" * 80)
    fecha_actual_utc = datetime.now(timezone.utc)
    print(f"\n[*] Fecha actual (UTC): {fecha_actual_utc.strftime('%Y-%m-%d %H:%M:%S')} UTC")
    print(f"[*] Buscando issues desde: hace 720 horas (30 días)")
    print(f"\n[*] Consulta JQL:")
    print(f"    {jql_query}\n")
    
    # Inicializar conexión a Jira
    try:
        print("[*] Conectando a Jira...")
        jira = JiraIntegration()
        print("[OK] Conexion establecida\n")
    except Exception as e:
        print(f"[ERROR] Error al conectar con Jira: {e}")
        return
    
    # Buscar issues - obtener todos los resultados disponibles
    try:
        if max_results and max_results > 0:
            print(f"[*] Buscando issues (máximo {max_results})...")
            issues = jira.search_issues(jql_query, max_results=max_results)
        else:
            print(f"[*] Buscando TODOS los issues disponibles...")
            # Pasar None para obtener todos los resultados
            issues = jira.search_issues(jql_query, max_results=None)
        print(f"\n[OK] Se encontraron {len(issues)} issues\n")
    except Exception as e:
        print(f"[ERROR] Error al buscar issues: {e}")
        return
    
    if not issues:
        print("[!] No se encontraron issues con los criterios especificados")
        return
    
    # Extraer claves
    claves = []
    for issue in issues:
        claves.append(issue.key)
    
    print(f"[*] Claves obtenidas: {len(claves)}")
    print(f"\n[*] Primeras 10 claves:")
    for i, clave in enumerate(claves[:10], 1):
        print(f"    {i}. {clave}")
    if len(claves) > 10:
        print(f"    ... y {len(claves) - 10} más")
    
    # Leer XLSX existente si existe para preservar otras columnas
    issues_existentes = {}
    columnas_existentes = ['Clave']
    
    if os.path.exists(archivo_xlsx):
        print(f"\n[*] Leyendo archivo existente: {archivo_xlsx}")
        try:
            wb = load_workbook(archivo_xlsx, data_only=True)
            ws = wb.active
            
            # Leer encabezados
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
                else:
                    headers.append('')
            
            columnas_existentes = headers.copy()
            
            # Leer datos existentes
            for row in ws.iter_rows(min_row=2, values_only=False):
                issue_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        col_name = headers[idx]
                        if col_name:
                            value = cell.value
                            issue_dict[col_name] = str(value) if value is not None else ''
                
                clave = issue_dict.get('Clave', '').strip()
                if clave:
                    issues_existentes[clave] = issue_dict
            
            print(f"[OK] Se leyeron {len(issues_existentes)} issues existentes")
        except Exception as e:
            print(f"[!] Error al leer archivo existente: {e}")
            print(f"    Se creará un archivo nuevo")
    
    # Crear estructura de datos con las nuevas claves
    # SIEMPRE crear registros nuevos (como si fuera la primera vez) - NO preservar datos existentes
    nuevos_issues = []
    for clave in claves:
        # Crear nuevo registro vacío siempre (como primera vez)
        nuevo_issue = {'Clave': clave}
        # Agregar otras columnas vacías si existen
        for col in columnas_existentes:
            if col != 'Clave' and col not in nuevo_issue:
                nuevo_issue[col] = ''
        nuevos_issues.append(nuevo_issue)
    
    # Si no hay columnas definidas, usar las estándar
    if len(columnas_existentes) == 1:  # Solo 'Clave'
        columnas_existentes = ['Clave', 'with RSOC', 'with Local Security', 'Closed', 'First response']
        # Agregar columnas faltantes a los issues existentes
        for issue in nuevos_issues:
            for col in columnas_existentes:
                if col not in issue:
                    issue[col] = ''
    
    # Guardar XLSX actualizado
    print(f"\n[*] Guardando {len(nuevos_issues)} issues en: {archivo_xlsx}")
    try:
        wb = Workbook()
        ws = wb.active
        
        # Escribir encabezados
        for col_idx, col_name in enumerate(columnas_existentes, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Escribir datos
        for row_idx, issue in enumerate(nuevos_issues, start=2):
            for col_idx, col_name in enumerate(columnas_existentes, start=1):
                value = issue.get(col_name, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(archivo_xlsx)
        print(f"[OK] Archivo guardado exitosamente")
        print(f"\n[*] Resumen:")
        print(f"    Total issues en XLSX: {len(nuevos_issues)}")
        print(f"    Issues actualizados desde Jira (todos como primera vez)")
        
    except Exception as e:
        print(f"[ERROR] Error al guardar el XLSX: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import sys
    
    # Permitir especificar el archivo XLSX como argumento
    archivo = sys.argv[1] if len(sys.argv) > 1 else "Libro1.xlsx"
    
    # Permitir especificar max_results como segundo argumento
    # Si no se especifica o es 0, obtener todos los resultados
    if len(sys.argv) > 2:
        max_results = int(sys.argv[2]) if int(sys.argv[2]) > 0 else None
    else:
        max_results = None  # Por defecto obtener todos
    
    obtener_issues_y_actualizar_xlsx(archivo, max_results)
    
    print("\n" + "=" * 80)
    print("[OK] Proceso completado")
    print("=" * 80)
    print("\n[*] Siguiente paso: Ejecutar 'python procesar_csv.py' para obtener las fechas")
