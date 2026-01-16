"""
Script para procesar XLSX y llenar fechas de cambio de estado
Lee Libro1.xlsx y busca fechas de cambio a "with RSOC" y "with Local Security"
"""
from jira_integration import JiraIntegration
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def parse_jira_date(date_str):
    """Convierte fecha de Jira a datetime"""
    if not date_str or (isinstance(date_str, str) and date_str.strip() == ''):
        return None
    
    try:
        # Si ya es un datetime, retornarlo
        if isinstance(date_str, datetime):
            return date_str
        # Formato: 2025-12-30T19:15:15.375-0500
        # Remover los milisegundos y el offset para simplificar
        date_part = str(date_str).split('.')[0]  # 2025-12-30T19:15:15
        return datetime.fromisoformat(date_part)
    except Exception as e:
        return None

def calcular_diferencias_horas(issue_data):
    """
    Calcula las diferencias en horas para las columnas I.First Response, I.Escalamiento, I.respuesta Sub
    """
    rsoc_str = issue_data.get('with RSOC', '')
    local_str = issue_data.get('with Local Security', '')
    closed_str = issue_data.get('Closed', '')
    first_response_str = issue_data.get('First response', '')
    
    rsoc_date = parse_jira_date(rsoc_str) if rsoc_str else None
    local_date = parse_jira_date(local_str) if local_str else None
    closed_date = parse_jira_date(closed_str) if closed_str else None
    first_response_date = parse_jira_date(first_response_str) if first_response_str else None
    
    # I.First Response: First response - with RSOC
    if first_response_date and rsoc_date:
        diferencia = first_response_date - rsoc_date
        horas = diferencia.total_seconds() / 3600
        issue_data['I.First Response'] = f"{horas:.2f}"
    else:
        issue_data['I.First Response'] = ''
    
    # I.Escalamiento: with Local Security - First response
    if local_date and first_response_date:
        diferencia = local_date - first_response_date
        horas = diferencia.total_seconds() / 3600
        issue_data['I.Escalamiento'] = f"{horas:.2f}"
    else:
        issue_data['I.Escalamiento'] = ''
    
    # I.respuesta Sub: Closed - First response
    if closed_date and first_response_date:
        diferencia = closed_date - first_response_date
        horas = diferencia.total_seconds() / 3600
        issue_data['I.respuesta Sub'] = f"{horas:.2f}"
    else:
        issue_data['I.respuesta Sub'] = ''

def procesar_csv(archivo_entrada='Libro1.xlsx', archivo_salida=None):
    """
    Procesa el XLSX y llena las columnas con fechas de cambio de estado
    
    Args:
        archivo_entrada: Nombre del archivo XLSX de entrada
        archivo_salida: Nombre del archivo XLSX de salida (si None, sobrescribe el original)
    """
    
    if archivo_salida is None:
        archivo_salida = archivo_entrada
    
    # Inicializar conexión a Jira
    try:
        print("[*] Conectando a Jira...")
        jira = JiraIntegration()
        print("[OK] Conexion establecida\n")
    except Exception as e:
        print(f"[ERROR] Error al conectar con Jira: {e}")
        return
    
    # Leer el XLSX
    print(f"[*] Leyendo archivo: {archivo_entrada}")
    issues = []
    
    try:
        wb = load_workbook(archivo_entrada, data_only=True)
        ws = wb.active
        
        # Leer encabezados
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else '')
        
        # Definir todas las columnas que necesitamos (asegurar que existan)
        todas_las_columnas = ['Clave', 'with RSOC', 'with Local Security', 'Closed', 'First response',
                             'I.First Response', 'I.Escalamiento', 'I.respuesta Sub']
        
        # Leer datos
        for row in ws.iter_rows(min_row=2, values_only=False):
            issue_dict = {}
            # Primero leer las columnas que existen en el Excel
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    col_name = headers[idx]
                    if col_name:
                        value = cell.value
                        # Convertir a string si no es None
                        issue_dict[col_name] = str(value) if value is not None else ''
            
            # Asegurar que todas las columnas necesarias existan en el diccionario
            for col_name in todas_las_columnas:
                if col_name not in issue_dict:
                    issue_dict[col_name] = ''
            
            if issue_dict.get('Clave', '').strip():
                issues.append(issue_dict)
        
        print(f"[OK] Se encontraron {len(issues)} issues en el XLSX\n")
    except Exception as e:
        print(f"[ERROR] Error al leer el XLSX: {e}")
        return
    
    if not issues:
        print("[ERROR] No se encontraron issues en el XLSX")
        return
    
    # PASO 1: Agregar fechas desde Jira (with RSOC, with Local Security, Closed, First response)
    print("\n" + "=" * 80)
    print("[PASO 1] Agregando fechas desde Jira...")
    print("=" * 80)
    print("-" * 80)
    
    encontrados_rsoc = 0
    encontrados_local = 0
    encontrados_closed = 0
    encontrados_first_response = 0
    errores = 0
    
    # Obtener lista de personas para First response desde config
    try:
        import config
        target_assignees = getattr(config, 'FIRST_RESPONSE_ASSIGNEES', [])
        if not target_assignees:
            print("[!] Advertencia: No se encontró FIRST_RESPONSE_ASSIGNEES en config.py")
            target_assignees = []
    except ImportError:
        # Si no hay config.py, intentar desde variables de entorno
        import os
        assignees_str = os.getenv('FIRST_RESPONSE_ASSIGNEES', '')
        if assignees_str:
            target_assignees = [a.strip() for a in assignees_str.split(',')]
        else:
            print("[!] Advertencia: No se encontró configuración de FIRST_RESPONSE_ASSIGNEES")
            target_assignees = []
    
    for i, issue_data in enumerate(issues, 1):
        issue_key = issue_data.get('Clave', '').strip()
        if not issue_key:
            continue
            
        print(f"[{i}/{len(issues)}] {issue_key}...", end=' ')
        
        try:
            # Buscar fecha de cambio a "with RSOC" - SIEMPRE buscar desde cero (como primera vez)
            rsoc_result = jira.get_status_change_date(issue_key, "with RSOC")
            if rsoc_result:
                fecha_rsoc = rsoc_result['date']
                issue_data['with RSOC'] = fecha_rsoc
                encontrados_rsoc += 1
                print(f"RSOC: OK ({fecha_rsoc})", end=' ')
            else:
                # Si no se encuentra, limpiar el valor (como primera vez)
                issue_data['with RSOC'] = ''
                print("RSOC: no encontrado", end=' ')
            
            # Buscar fecha de cambio a "with Local Security" - SIEMPRE buscar desde cero (como primera vez)
            local_result = jira.get_status_change_date(issue_key, "with Local Security")
            if local_result:
                fecha_local = local_result['date']
                issue_data['with Local Security'] = fecha_local
                encontrados_local += 1
                print(f"Local: OK ({fecha_local})", end=' ')
            else:
                # Si no se encuentra, limpiar el valor (como primera vez)
                issue_data['with Local Security'] = ''
                print("Local: no encontrado", end=' ')
            
            # Buscar fecha de cambio a "Closed" - SIEMPRE buscar desde cero (como primera vez)
            closed_result = jira.get_status_change_date(issue_key, "Closed")
            if closed_result:
                fecha_closed = closed_result['date']
                issue_data['Closed'] = fecha_closed
                encontrados_closed += 1
                print(f"Closed: OK ({fecha_closed})", end=' ')
            else:
                # Si no se encuentra, limpiar el valor (como primera vez)
                issue_data['Closed'] = ''
                print("Closed: no encontrado", end=' ')
            
            # Buscar fecha de asignación a personas específicas (First response) - SIEMPRE buscar desde cero (como primera vez)
            first_response_result = jira.get_assignee_change_date(issue_key, target_assignees)
            if first_response_result:
                fecha_first = first_response_result['date']
                issue_data['First response'] = fecha_first
                encontrados_first_response += 1
                print(f"First response: OK ({fecha_first})", end=' ')
            else:
                # Si no se encuentra, limpiar el valor (como primera vez)
                issue_data['First response'] = ''
                print("First response: no encontrado", end=' ')
            
            print()  # Nueva línea
            
        except Exception as e:
            errores += 1
            print(f"[ERROR] {e}")
            # Continuar con el siguiente issue
        
        # Mostrar progreso cada 10 issues
        if i % 10 == 0:
            print(f"\n[*] Progreso: {i}/{len(issues)} procesados")
            print(f"    RSOC encontrados: {encontrados_rsoc}")
            print(f"    Local Security encontrados: {encontrados_local}")
            print(f"    Closed encontrados: {encontrados_closed}")
            print(f"    First response encontrados: {encontrados_first_response}")
            print(f"    Errores: {errores}\n")
    
    print("-" * 80)
    print(f"\n[OK] Paso 1 completado - Fechas agregadas")
    print(f"    Total issues: {len(issues)}")
    print(f"    RSOC encontrados: {encontrados_rsoc}")
    print(f"    Local Security encontrados: {encontrados_local}")
    print(f"    Closed encontrados: {encontrados_closed}")
    print(f"    First response encontrados: {encontrados_first_response}")
    print(f"    Errores: {errores}")
    
    # PASO 2: Calcular diferencias en horas (I.First Response, I.Escalamiento, I.respuesta Sub)
    print("\n" + "=" * 80)
    print("[PASO 2] Calculando diferencias en horas...")
    print("=" * 80)
    calculados_diferencias = 0
    for issue_data in issues:
        calcular_diferencias_horas(issue_data)
        if issue_data.get('I.First Response') or issue_data.get('I.Escalamiento') or issue_data.get('I.respuesta Sub'):
            calculados_diferencias += 1
    print(f"[OK] Paso 2 completado - Diferencias calculadas para {calculados_diferencias} issues")
    
    # PASO 3: Filtrar y eliminar filas cuando Escalamiento (with Local Security) < First response
    print("\n" + "=" * 80)
    print("[PASO 3] Filtrando filas donde Escalamiento < First response...")
    print("=" * 80)
    issues_originales = len(issues)
    issues_filtrados = []
    eliminados = 0
    
    for issue_data in issues:
        escalamiento_str = issue_data.get('with Local Security', '').strip()
        first_response_str = issue_data.get('First response', '').strip()
        
        # Solo eliminar si ambas fechas existen
        if escalamiento_str and first_response_str:
            try:
                # Convertir las fechas a datetime para comparar
                escalamiento_date = parse_jira_date(escalamiento_str)
                first_response_date = parse_jira_date(first_response_str)
                
                # Eliminar solo si Escalamiento < First response
                if escalamiento_date and first_response_date and escalamiento_date < first_response_date:
                    eliminados += 1
                    print(f"    [ELIMINADO] {issue_data.get('Clave', 'N/A')}: Escalamiento ({escalamiento_date}) < First response ({first_response_date})")
                    continue  # Saltar este issue, no agregarlo a issues_filtrados
            except Exception as e:
                # Si hay error al parsear fechas, mantener la fila
                pass
        
        # Si no cumple la condición o no tiene ambas fechas, mantenerlo
        issues_filtrados.append(issue_data)
    
    issues = issues_filtrados  # Reemplazar con la lista filtrada
    print(f"[OK] Paso 3 completado - Filtrado: {eliminados} fila(s) eliminada(s) de {issues_originales} totales")
    print(f"    Issues restantes: {len(issues)}")
    
    # PASO 4: Escribir el XLSX actualizado con todas las columnas
    print("\n" + "=" * 80)
    print(f"[PASO 4] Guardando resultados en: {archivo_salida}")
    print("=" * 80)
    
    # Definir todas las columnas que deben estar en el Excel final
    fieldnames = ['Clave', 'with RSOC', 'with Local Security', 'Closed', 'First response',
                 'I.First Response', 'I.Escalamiento', 'I.respuesta Sub']
    
    # Asegurar que TODOS los issues tengan TODAS las columnas antes de guardar
    print(f"[*] Verificando que todos los issues tengan todas las columnas...")
    for issue_data in issues:
        for col_name in fieldnames:
            if col_name not in issue_data:
                issue_data[col_name] = ''
    print(f"[OK] Todos los issues tienen todas las columnas ({len(fieldnames)} columnas)")
    
    try:
        wb = Workbook()
        ws = wb.active
        
        # Escribir encabezados - VERIFICAR que se escriban todos
        print(f"[DEBUG] Escribiendo {len(fieldnames)} encabezados...")
        for col_idx, col_name in enumerate(fieldnames, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
            print(f"[DEBUG] Encabezado {col_idx}: {col_name}")
        
        # Escribir datos - TODAS las columnas para TODOS los issues
        print(f"[*] Escribiendo {len(issues)} issues con {len(fieldnames)} columnas...")
        print(f"[DEBUG] Columnas a escribir: {fieldnames}")
        
        for row_idx, issue_data in enumerate(issues, start=2):
            # Debug: mostrar las primeras 3 issues
            if row_idx <= 4:
                print(f"[DEBUG] Issue {row_idx-1}: Clave={issue_data.get('Clave', 'N/A')}, tiene {len(issue_data)} columnas")
                print(f"         Columnas: {list(issue_data.keys())}")
            
            for col_idx, col_name in enumerate(fieldnames, start=1):
                # Obtener el valor del diccionario
                value = issue_data.get(col_name, '')
                # Debug para las primeras filas
                if row_idx <= 4 and col_name in ['with RSOC', 'with Local Security', 'First response']:
                    print(f"[DEBUG] Escribiendo {col_name} = '{value}' en fila {row_idx}, columna {col_idx}")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(archivo_salida)
        print(f"[OK] Archivo guardado exitosamente: {archivo_salida}")
        print(f"[DEBUG] Verificando archivo guardado...")
        # Verificar que el archivo se guardó correctamente
        try:
            wb_check = load_workbook(archivo_salida, data_only=True)
            ws_check = wb_check.active
            headers_check = [cell.value for cell in ws_check[1]]
            print(f"[DEBUG] Columnas en archivo guardado: {headers_check}")
            print(f"[DEBUG] Total columnas: {len(headers_check)}")
        except Exception as e:
            print(f"[DEBUG] Error al verificar archivo: {e}")
        
        # Estadísticas finales
        con_rsoc = sum(1 for i in issues if i.get('with RSOC', '').strip())
        con_local = sum(1 for i in issues if i.get('with Local Security', '').strip())
        con_closed = sum(1 for i in issues if i.get('Closed', '').strip())
        con_first_response = sum(1 for i in issues if i.get('First response', '').strip())
        completos = sum(1 for i in issues if i.get('with RSOC', '').strip() and i.get('with Local Security', '').strip())
        
        print(f"\n[*] Estadisticas finales:")
        print(f"    Issues con fecha 'with RSOC': {con_rsoc}")
        print(f"    Issues con fecha 'with Local Security': {con_local}")
        print(f"    Issues con fecha 'Closed': {con_closed}")
        print(f"    Issues con fecha 'First response': {con_first_response}")
        print(f"    Issues completos (ambas fechas): {completos}")
        
    except Exception as e:
        print(f"[ERROR] Error al guardar el XLSX: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # Procesar el XLSX
    # Por defecto sobrescribe el archivo original, pero puedes crear una copia primero
    import sys
    
    archivo_entrada = sys.argv[1] if len(sys.argv) > 1 else "Libro1.xlsx"
    
    if not os.path.exists(archivo_entrada):
        print(f"[ERROR] El archivo {archivo_entrada} no existe")
        sys.exit(1)
    
    # Crear copia de respaldo
    if os.path.exists(archivo_entrada):
        import shutil
        backup = f"{archivo_entrada}.backup"
        shutil.copy2(archivo_entrada, backup)
        print(f"[*] Copia de respaldo creada: {backup}\n")
    
    procesar_csv(archivo_entrada)
